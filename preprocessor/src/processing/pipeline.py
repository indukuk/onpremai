"""Main processing pipeline for file ingestion.

Orchestrates the full workflow: detect file type -> extract content ->
generate schema -> detect joins -> write metadata.json -> notify memory.

This is the central coordination point that the triggers invoke.
"""

from __future__ import annotations

import io
import json
import os
from typing import Any

import structlog

from src.config import PreprocessorSettings, OcrBackend, PdfBackend
from src.extraction.base import ExtractionBackend, ExtractionResult
from src.extraction.pdfplumber_ext import PdfplumberBackend
from src.extraction.tesseract import TesseractBackend
from src.extraction.textract import TextractBackend
from src.extraction.tika import TikaBackend
from src.health import record_processing
from src.idempotency import IdempotencyTracker
from src.models import (
    EvidenceType,
    FileMetadata,
    FileType,
    ProcessingError,
    ProcessingResult,
    SheetSchema,
)
from src.processing.join_detector import detect_joins
from src.processing.schema_detector import SAMPLE_ROW_COUNT, detect_schema

logger = structlog.get_logger(__name__)

# File extension to FileType mapping
_EXTENSION_MAP: dict[str, FileType] = {
    ".xlsx": FileType.EXCEL,
    ".xls": FileType.EXCEL,
    ".csv": FileType.CSV,
    ".tsv": FileType.CSV,
    ".pdf": FileType.PDF,
    ".docx": FileType.WORD,
    ".doc": FileType.WORD,
    ".pptx": FileType.POWERPOINT,
    ".ppt": FileType.POWERPOINT,
    ".png": FileType.IMAGE,
    ".jpg": FileType.IMAGE,
    ".jpeg": FileType.IMAGE,
    ".tiff": FileType.IMAGE,
    ".tif": FileType.IMAGE,
    ".bmp": FileType.IMAGE,
    ".json": FileType.JSON,
}


class ProcessingPipeline:
    """Orchestrates the full file processing workflow.

    Coordinates file type detection, content extraction, schema generation,
    join detection, metadata writing, and memory notification.
    """

    def __init__(
        self,
        settings: PreprocessorSettings,
        tracker: IdempotencyTracker,
    ) -> None:
        """Initialize the pipeline with configuration and tracker.

        Args:
            settings: Service configuration.
            tracker: Idempotency tracker for skip decisions.
        """
        self._settings = settings
        self._tracker = tracker
        self._ocr_backend = self._create_ocr_backend()
        self._pdf_backend = self._create_pdf_backend()
        self._doc_backend = TikaBackend()

    def _create_ocr_backend(self) -> ExtractionBackend:
        """Create the configured OCR backend instance."""
        if self._settings.ocr_backend == OcrBackend.TEXTRACT:
            return TextractBackend()
        return TesseractBackend()

    def _create_pdf_backend(self) -> ExtractionBackend:
        """Create the configured PDF extraction backend instance."""
        if self._settings.pdf_backend == PdfBackend.PDFPLUMBER:
            return PdfplumberBackend()
        # For pymupdf and tika, fall back to pdfplumber as default
        return PdfplumberBackend()

    async def process_file(self, file_key: str) -> ProcessingResult:
        """Process a single file through the full pipeline.

        Steps:
        1. Download file from storage
        2. Check idempotency (skip if unchanged)
        3. Detect file type
        4. Extract content
        5. Generate schema (structured files)
        6. Detect joins (if other structured files exist)
        7. Write metadata.json
        8. Notify memory service

        Args:
            file_key: Storage key of the file to process.

        Returns:
            ProcessingResult indicating success, skip, or failure.
        """
        from common.clients import StorageClient

        storage = StorageClient(
            backend=self._settings.storage_backend,
            bucket=self._settings.storage_bucket,
            endpoint=self._settings.storage_endpoint,
            access_key=self._settings.storage_access_key,
            secret_key=self._settings.storage_secret_key,
        )

        try:
            return await self._process_file_impl(file_key, storage)
        finally:
            await storage.close()

    async def _process_file_impl(
        self, file_key: str, storage: Any
    ) -> ProcessingResult:
        """Internal implementation of file processing."""
        # Step 1: Download file
        try:
            file_bytes = await storage.get(file_key)
        except Exception as exc:
            logger.error("file_download_failed", file_key=file_key, error=str(exc))
            record_processing(success=False)
            return ProcessingResult(
                success=False,
                file_key=file_key,
                error_message=f"Failed to download file: {exc}",
            )

        # Step 2: Size check
        if len(file_bytes) > self._settings.max_file_size_bytes:
            logger.warning(
                "file_too_large",
                file_key=file_key,
                size_bytes=len(file_bytes),
                max_bytes=self._settings.max_file_size_bytes,
            )
            error_metadata = self._build_error_metadata(
                file_key=file_key,
                file_bytes=file_bytes,
                error_type="file_too_large",
                message=f"File exceeds maximum size of {self._settings.max_file_size_mb}MB",
            )
            await self._write_metadata(storage, file_key, error_metadata)
            record_processing(success=False)
            return ProcessingResult(
                success=False,
                file_key=file_key,
                metadata=error_metadata,
                error_message="File too large",
            )

        # Step 3: Idempotency check
        content_hash = IdempotencyTracker.compute_hash(file_bytes)

        # Check metadata.json in storage for durability across restarts
        metadata_key = self._metadata_key(file_key)
        try:
            if await storage.exists(metadata_key):
                existing_meta = await storage.get_json(metadata_key)
                if existing_meta.get("content_hash") == content_hash:
                    self._tracker.mark_from_metadata(file_key, existing_meta)
                    logger.debug("file_unchanged_skipped", file_key=file_key)
                    return ProcessingResult(
                        success=True, file_key=file_key, skipped=True
                    )
        except Exception:
            # If we cannot check, proceed with processing
            pass

        if self._tracker.is_processed(file_key, content_hash):
            logger.debug("file_already_processed", file_key=file_key)
            return ProcessingResult(success=True, file_key=file_key, skipped=True)

        # Step 4: Detect file type
        file_type = self._detect_file_type(file_key, file_bytes)
        if file_type == FileType.UNKNOWN:
            logger.info("unsupported_file_type", file_key=file_key)
            return ProcessingResult(
                success=True,
                file_key=file_key,
                skipped=True,
                error_message="Unsupported file type",
            )

        logger.info(
            "processing_file",
            file_key=file_key,
            file_type=file_type.value,
            size_bytes=len(file_bytes),
        )

        # Step 5: Extract content based on file type
        metadata = await self._extract_and_build_metadata(
            file_key=file_key,
            file_bytes=file_bytes,
            file_type=file_type,
            content_hash=content_hash,
            storage=storage,
        )

        # Step 6: Write metadata.json
        await self._write_metadata(storage, file_key, metadata)

        # Step 7: Mark as processed
        self._tracker.mark_processed(file_key, content_hash)

        # Step 8: Notify memory service (best-effort)
        await self._notify_memory(metadata)

        # Step 9: Publish event for shadow agent notification (best-effort)
        await self._publish_ingestion_event(metadata)

        record_processing(success=True)
        logger.info(
            "file_processed_successfully",
            file_key=file_key,
            file_type=file_type.value,
            evidence_type=metadata.evidence_type.value,
        )

        return ProcessingResult(
            success=True,
            file_key=file_key,
            metadata=metadata,
        )

    async def _extract_and_build_metadata(
        self,
        file_key: str,
        file_bytes: bytes,
        file_type: FileType,
        content_hash: str,
        storage: Any,
    ) -> FileMetadata:
        """Route to the appropriate processor and build metadata."""
        filename = os.path.basename(file_key)
        tenant_id = self._extract_tenant_id(file_key)

        base_metadata = FileMetadata(
            filename=filename,
            file_key=file_key,
            file_type=file_type,
            size_bytes=len(file_bytes),
            content_hash=content_hash,
            tenant_id=tenant_id,
        )

        try:
            if file_type in (FileType.EXCEL, FileType.CSV):
                return await self._process_structured(
                    file_bytes, file_type, base_metadata, storage
                )
            elif file_type == FileType.PDF:
                return await self._process_pdf(file_bytes, base_metadata)
            elif file_type in (FileType.WORD, FileType.POWERPOINT):
                return await self._process_document(file_bytes, file_type, base_metadata)
            elif file_type == FileType.IMAGE:
                return await self._process_image(file_bytes, base_metadata)
            elif file_type == FileType.JSON:
                return await self._process_json(file_bytes, base_metadata)
            else:
                return base_metadata
        except Exception as exc:
            logger.error(
                "extraction_failed",
                file_key=file_key,
                error=str(exc),
                error_type=type(exc).__name__,
            )
            base_metadata.error = ProcessingError(
                error_type="extraction_failed",
                message=str(exc),
                recoverable=True,
            )
            return base_metadata

    async def _process_structured(
        self,
        file_bytes: bytes,
        file_type: FileType,
        metadata: FileMetadata,
        storage: Any,
    ) -> FileMetadata:
        """Process Excel or CSV files into structured schema."""
        if file_type == FileType.EXCEL:
            sheets = self._parse_excel(file_bytes)
        else:
            sheets = self._parse_csv(file_bytes)

        metadata.sheets = sheets
        metadata.evidence_type = EvidenceType.STRUCTURED

        # Detect joins with other files in the same prefix
        if sheets:
            all_columns = []
            for sheet in sheets:
                all_columns.extend(sheet.columns)

            other_files = await self._get_sibling_metadata(
                metadata.file_key, storage
            )
            if other_files:
                metadata.join_candidates = detect_joins(
                    metadata.file_key, all_columns, other_files
                )

        return metadata

    async def _process_pdf(
        self, file_bytes: bytes, metadata: FileMetadata
    ) -> FileMetadata:
        """Process PDF - try digital extraction first, fall back to OCR."""
        # Try digital PDF extraction first
        result = await self._pdf_backend.extract_text(file_bytes)

        if result.success and result.text.strip():
            # Check if it looks like a scanned PDF (very little text per page)
            needs_ocr = result.metadata.get("needs_ocr", False)
            if not needs_ocr:
                metadata.extracted_text = result.text
                metadata.page_count = result.page_count
                metadata.evidence_type = EvidenceType.UNSTRUCTURED
                return metadata

        # Fall back to OCR for scanned PDFs
        if self._ocr_backend.is_available():
            ocr_result = await self._ocr_backend.extract_text(file_bytes)
            if ocr_result.success:
                metadata.extracted_text = ocr_result.text
                metadata.page_count = ocr_result.page_count or result.page_count
                metadata.evidence_type = EvidenceType.UNSTRUCTURED
                return metadata
            logger.warning(
                "ocr_fallback_failed",
                file_key=metadata.file_key,
                error=ocr_result.error_message,
            )

        # Use whatever we got from digital extraction
        metadata.extracted_text = result.text if result.success else ""
        metadata.page_count = result.page_count
        metadata.evidence_type = EvidenceType.UNSTRUCTURED

        if not result.success:
            metadata.error = ProcessingError(
                error_type="extraction_failed",
                message=result.error_message,
                recoverable=True,
            )

        return metadata

    async def _process_document(
        self, file_bytes: bytes, file_type: FileType, metadata: FileMetadata
    ) -> FileMetadata:
        """Process Word or PowerPoint documents."""
        kwargs: dict[str, object] = {}
        if file_type == FileType.POWERPOINT:
            kwargs["file_type"] = "pptx"
        else:
            kwargs["file_type"] = "docx"

        result = await self._doc_backend.extract_text(file_bytes, **kwargs)

        if result.success:
            metadata.extracted_text = result.text
            metadata.page_count = result.page_count
        else:
            metadata.error = ProcessingError(
                error_type="extraction_failed",
                message=result.error_message,
                recoverable=True,
            )

        metadata.evidence_type = EvidenceType.UNSTRUCTURED
        return metadata

    async def _process_image(
        self, file_bytes: bytes, metadata: FileMetadata
    ) -> FileMetadata:
        """Process images via OCR backend."""
        if not self._ocr_backend.is_available():
            metadata.error = ProcessingError(
                error_type="backend_unavailable",
                message=f"OCR backend '{self._ocr_backend.name}' is not available",
                recoverable=False,
            )
            metadata.evidence_type = EvidenceType.UNSTRUCTURED
            return metadata

        result = await self._ocr_backend.extract_text(file_bytes)

        if result.success:
            metadata.extracted_text = result.text
            metadata.page_count = 1
        else:
            metadata.error = ProcessingError(
                error_type="ocr_failed",
                message=result.error_message,
                recoverable=True,
            )

        metadata.evidence_type = EvidenceType.UNSTRUCTURED
        return metadata

    async def _process_json(
        self, file_bytes: bytes, metadata: FileMetadata
    ) -> FileMetadata:
        """Process JSON files - parse and infer schema."""
        try:
            content = json.loads(file_bytes.decode("utf-8"))
        except (json.JSONDecodeError, UnicodeDecodeError) as exc:
            metadata.error = ProcessingError(
                error_type="parse_error",
                message=f"Invalid JSON: {exc}",
                recoverable=False,
            )
            return metadata

        if isinstance(content, list) and content and isinstance(content[0], dict):
            # Array of objects - treat as tabular
            columns = list(content[0].keys())
            rows = content[:1000]  # Cap for schema detection
            schema = detect_schema(columns, rows, sheet_name="root")
            metadata.sheets = [schema]
            metadata.evidence_type = EvidenceType.STRUCTURED
        elif isinstance(content, dict):
            # Single object - extract keys as schema hint
            metadata.extracted_text = json.dumps(content, indent=2, default=str)[:10000]
            metadata.evidence_type = EvidenceType.MIXED
        else:
            metadata.extracted_text = json.dumps(content, default=str)[:10000]
            metadata.evidence_type = EvidenceType.UNSTRUCTURED

        return metadata

    def _parse_excel(self, file_bytes: bytes) -> list[SheetSchema]:
        """Parse Excel file and detect schema for each sheet."""
        import openpyxl

        sheets: list[SheetSchema] = []

        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        try:
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows_iter = ws.iter_rows(values_only=True)

                # First row is headers
                try:
                    header_row = next(rows_iter)
                except StopIteration:
                    continue

                columns = [str(cell) if cell is not None else f"column_{i}" for i, cell in enumerate(header_row)]

                # Read all data rows
                data_rows: list[dict[str, Any]] = []
                for row_values in rows_iter:
                    row_dict: dict[str, Any] = {}
                    for i, val in enumerate(row_values):
                        if i < len(columns):
                            row_dict[columns[i]] = val
                    data_rows.append(row_dict)

                schema = detect_schema(columns, data_rows, sheet_name=sheet_name)
                sheets.append(schema)
        finally:
            wb.close()

        return sheets

    def _parse_csv(self, file_bytes: bytes) -> list[SheetSchema]:
        """Parse CSV/TSV file and detect schema."""
        import csv

        try:
            text = file_bytes.decode("utf-8")
        except UnicodeDecodeError:
            text = file_bytes.decode("latin-1")

        # Detect delimiter
        sniffer = csv.Sniffer()
        try:
            dialect = sniffer.sniff(text[:4096])
            delimiter = dialect.delimiter
        except csv.Error:
            delimiter = "," if "," in text[:1000] else "\t"

        reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
        columns = reader.fieldnames or []

        data_rows: list[dict[str, Any]] = []
        for row in reader:
            data_rows.append(dict(row))

        schema = detect_schema(list(columns), data_rows, sheet_name="data")
        return [schema]

    async def _get_sibling_metadata(
        self, file_key: str, storage: Any
    ) -> list[dict[str, Any]]:
        """Get metadata from other processed files in the same prefix.

        Used for join detection across files in the same tenant/control prefix.
        """
        # Determine the parent prefix (one level up from the file)
        parts = file_key.rsplit("/", 1)
        if len(parts) < 2:
            return []
        prefix = parts[0] + "/"

        try:
            keys = await storage.list_objects(prefix)
        except Exception:
            return []

        metadata_files: list[dict[str, Any]] = []
        for key in keys:
            if key.endswith("metadata.json") and key != self._metadata_key(file_key):
                try:
                    meta = await storage.get_json(key)
                    metadata_files.append(meta)
                except Exception:
                    continue

        return metadata_files

    async def _write_metadata(
        self, storage: Any, file_key: str, metadata: FileMetadata
    ) -> None:
        """Write metadata.json to storage alongside the original file."""
        metadata_key = self._metadata_key(file_key)
        try:
            await storage.put_json(metadata_key, metadata.model_dump(mode="json"))
        except Exception as exc:
            logger.error(
                "metadata_write_failed",
                file_key=file_key,
                metadata_key=metadata_key,
                error=str(exc),
            )

    async def _notify_memory(self, metadata: FileMetadata) -> None:
        """Notify memory service about newly processed file (best-effort)."""
        if not self._settings.memory_url:
            return

        if not metadata.tenant_id:
            return

        try:
            from common.clients import MemoryClient

            memory = MemoryClient(memory_url=self._settings.memory_url)
            try:
                await memory.tenant_store(
                    tenant_id=metadata.tenant_id,
                    fact=f"New file uploaded: {metadata.filename}, type: {metadata.file_type.value}",
                    category="evidence",
                    metadata={
                        "source": "preprocessor",
                        "file_key": metadata.file_key,
                        "evidence_type": metadata.evidence_type.value,
                    },
                )
            finally:
                await memory.close()
        except Exception as exc:
            logger.warning(
                "memory_notification_failed",
                file_key=metadata.file_key,
                error=str(exc),
            )

    async def _publish_ingestion_event(self, metadata: FileMetadata) -> None:
        """Push an event to the memory-service event queue for shadow agent notification.

        Best-effort: failures are logged as warnings and never crash file processing.
        """
        if not self._settings.memory_url:
            return

        if not metadata.tenant_id:
            return

        try:
            from common.clients import MemoryClient

            memory = MemoryClient(memory_url=self._settings.memory_url)
            try:
                await memory.event_queue_push(
                    user_id="__all__",
                    tenant_id=metadata.tenant_id,
                    event_type="evidence_uploaded",
                    summary=f"New evidence uploaded: {metadata.filename} ({metadata.file_type.value})",
                    priority="low",
                    source_service="preprocessor",
                    metadata={
                        "filename": metadata.filename,
                        "file_type": metadata.file_type.value,
                        "storage_key": metadata.file_key,
                    },
                )
            finally:
                await memory.close()
        except Exception as exc:
            logger.warning(
                "event_publish_failed",
                file_key=metadata.file_key,
                error=str(exc),
            )

    def _detect_file_type(self, file_key: str, file_bytes: bytes) -> FileType:
        """Detect file type from extension and optional magic bytes.

        Args:
            file_key: Storage key (used for extension detection).
            file_bytes: File content (for magic byte verification).

        Returns:
            Detected FileType or UNKNOWN.
        """
        # Extension-based detection
        ext = os.path.splitext(file_key)[1].lower()
        file_type = _EXTENSION_MAP.get(ext, FileType.UNKNOWN)

        if file_type != FileType.UNKNOWN:
            return file_type

        # Try magic bytes as fallback
        try:
            import magic

            mime = magic.from_buffer(file_bytes[:2048], mime=True)
            return self._mime_to_file_type(mime)
        except (ImportError, Exception):
            return FileType.UNKNOWN

    @staticmethod
    def _mime_to_file_type(mime: str) -> FileType:
        """Convert MIME type to FileType enum."""
        mime_map: dict[str, FileType] = {
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": FileType.EXCEL,
            "application/vnd.ms-excel": FileType.EXCEL,
            "text/csv": FileType.CSV,
            "application/pdf": FileType.PDF,
            "application/vnd.openxmlformats-officedocument.wordprocessingml.document": FileType.WORD,
            "application/msword": FileType.WORD,
            "application/vnd.openxmlformats-officedocument.presentationml.presentation": FileType.POWERPOINT,
            "application/json": FileType.JSON,
            "image/png": FileType.IMAGE,
            "image/jpeg": FileType.IMAGE,
            "image/tiff": FileType.IMAGE,
        }
        return mime_map.get(mime, FileType.UNKNOWN)

    @staticmethod
    def _metadata_key(file_key: str) -> str:
        """Compute the metadata.json key for a given file key.

        Places metadata.json alongside the original file with a predictable name.
        Example: uploads/tenant-1/file.xlsx -> uploads/tenant-1/file.xlsx.metadata.json
        """
        return f"{file_key}.metadata.json"

    @staticmethod
    def _extract_tenant_id(file_key: str) -> str:
        """Extract tenant ID from file path convention.

        Assumes path format: {prefix}/{tenant_id}/...
        Example: uploads/tenant-123/evidence/file.xlsx -> tenant-123
        """
        parts = file_key.split("/")
        # Skip the watch prefix and take the next segment as tenant ID
        if len(parts) >= 2:
            # If first part looks like a prefix (uploads, etc.), take second
            if parts[0] in ("uploads", "evidence", "data"):
                return parts[1] if len(parts) > 1 else ""
            return parts[0]
        return ""
